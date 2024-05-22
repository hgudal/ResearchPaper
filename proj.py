import itertools
import openreview
import openreview.tools
import random
import string
from collections import Counter
import pandas as pd
from sentence_transformers import SentenceTransformer
import numpy as np
from scipy.optimize import linear_sum_assignment
from scipy.optimize import linear_sum_assignment
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize the OpenReview client and SBERT model
client = openreview.api.OpenReviewClient(
    baseurl='https://api2.openreview.net',
    username='hassangudal@umass.edu',
    password='Somaliland1!'
)
model = SentenceTransformer('all-mpnet-base-v2')

# Define research areas
research_areas = [
    "Artificial intelligence", "Computer vision", "Machine learning", "Natural language processing", 
    "The Web & information retrieval", "Computer architecture", "Computer networks", "Computer security", 
    "Databases", "Design automation", "Embedded & real-time systems", "High-performance computing", 
    "Mobile computing", "Measurement & performance analysis", "Operating systems", "Programming languages", 
    "Software engineering", "Algorithms & complexity", "Cryptography", "Logic & verification", 
    "Computational biology & bioinformatics", "Computer graphics", "Computer science education", "Economics & computation", 
    "Human-computer interaction", "Robotics", "Visualization"
]

university_data = {
    'Carnegie Mellon University': {
        'Artificial intelligence': 244, 'Computer vision': 385, 'Machine learning': 835, 'Natural language processing': 424,
        'The Web & information retrieval': 76, 'Computer architecture': 86, 'Computer networks': 85, 'Computer security': 126,
        'Databases': 93, 'Design automation': 18, 'Embedded & real-time systems': 22, 'High-performance computing': 10,
        'Mobile computing': 28, 'Measurement & performance analysis': 30, 'Operating systems': 53, 'Programming languages': 59,
        'Software engineering': 105, 'Algorithms & complexity': 192, 'Cryptography': 49, 'Logic & verification': 17,
        'Computational biology & bioinformatics': 29, 'Computer graphics': 98, 'Computer science education': 3, 'Economics & computation': 26,
        'Human-computer interaction': 505, 'Robotics': 340, 'Visualization': 10
    },
    'University of Illinois at Urbana-Champaign': {
        'Artificial intelligence': 116, 'Computer vision': 214, 'Machine learning': 361, 'Natural language processing': 204,
        'The Web & information retrieval': 124, 'Computer architecture': 201, 'Computer networks': 53, 'Computer security': 131,
        'Databases': 46, 'Design automation': 53, 'Embedded & real-time systems': 33, 'High-performance computing': 56,
        'Mobile computing': 40, 'Measurement & performance analysis': 24, 'Operating systems': 25, 'Programming languages': 45,
        'Software engineering': 69, 'Algorithms & complexity': 93, 'Cryptography': 29, 'Logic & verification': 35,
        'Computational biology & bioinformatics': 18, 'Computer graphics': 1, 'Computer science education': 10, 'Economics & computation': 8,
        'Human-computer interaction': 105, 'Robotics': 114, 'Visualization': 11
    },
    'University of California - San Diego': {
        'Artificial intelligence': 84, 'Computer vision': 309, 'Machine learning': 408, 'Natural language processing': 168,
        'The Web & information retrieval': 23, 'Computer architecture': 70, 'Computer networks': 54, 'Computer security': 108,
        'Databases': 62, 'Design automation': 132, 'Embedded & real-time systems': 1, 'High-performance computing': 9,
        'Mobile computing': 70, 'Measurement & performance analysis': 61, 'Operating systems': 18, 'Programming languages': 41,
        'Software engineering': 15, 'Algorithms & complexity': 84, 'Cryptography': 34, 'Logic & verification': 5,
        'Computational biology & bioinformatics': 16, 'Computer graphics': 84, 'Computer science education': 51, 'Economics & computation': 2,
        'Human-computer interaction': 149, 'Robotics': 149, 'Visualization': 5
    },
    'Georgia Institute of Technology': {
        'Artificial intelligence': 108, 'Computer vision': 179, 'Machine learning': 340, 'Natural language processing': 82,
        'The Web & information retrieval': 50, 'Computer architecture': 90, 'Computer networks': 19, 'Computer security': 258,
        'Databases': 58, 'Design automation': 47, 'Embedded & real-time systems': 1, 'High-performance computing': 47,
        'Mobile computing': 14, 'Measurement & performance analysis': 50, 'Operating systems': 17, 'Programming languages': 15,
        'Software engineering': 16, 'Algorithms & complexity': 72, 'Cryptography': 25, 'Logic & verification': 6,
        'Computational biology & bioinformatics': 7, 'Computer graphics': 30, 'Computer science education': 8, 'Economics & computation': 10,
        'Human-computer interaction': 210, 'Robotics': 302, 'Visualization': 49
    },
    'Massachusetts Institute of Technology': {
        'Artificial intelligence': 104, 'Computer vision': 247, 'Machine learning': 646, 'Natural language processing': 82,
        'The Web & information retrieval': 1, 'Computer architecture': 100, 'Computer networks': 108, 'Computer security': 31,
        'Databases': 82, 'Design automation': 10, 'Embedded & real-time systems': 0, 'High-performance computing': 1,
        'Mobile computing': 22, 'Measurement & performance analysis': 15, 'Operating systems': 62, 'Programming languages': 52,
        'Software engineering': 15, 'Algorithms & complexity': 197, 'Cryptography': 27, 'Logic & verification': 11,
        'Computational biology & bioinformatics': 11, 'Computer graphics': 120, 'Computer science education': 0, 'Economics & computation': 8,
        'Human-computer interaction': 93, 'Robotics': 315, 'Visualization': 0
    },
    'University of California - Berkeley': {
        'Artificial intelligence': 82, 'Computer vision': 360, 'Machine learning': 922, 'Natural language processing': 139,
        'The Web & information retrieval': 4, 'Computer architecture': 26, 'Computer networks': 99, 'Computer security': 105,
        'Databases': 119, 'Design automation': 33, 'Embedded & real-time systems': 12, 'High-performance computing': 17,
        'Mobile computing': 17, 'Measurement & performance analysis': 11, 'Operating systems': 57, 'Programming languages': 11,
        'Software engineering': 26, 'Algorithms & complexity': 124, 'Cryptography': 37, 'Logic & verification': 18,
        'Computational biology & bioinformatics': 7, 'Computer graphics': 29, 'Computer science education': 1, 'Economics & computation': 1,
        'Human-computer interaction': 92, 'Robotics': 413, 'Visualization': 5
    },
    'University of Michigan': {
        'Artificial intelligence': 74, 'Computer vision': 117, 'Machine learning': 144, 'Natural language processing': 139,
        'The Web & information retrieval': 48, 'Computer architecture': 182, 'Computer networks': 27, 'Computer security': 90,
        'Databases': 62, 'Design automation': 49, 'Embedded & real-time systems': 15, 'High-performance computing': 7,
        'Mobile computing': 37, 'Measurement & performance analysis': 29, 'Operating systems': 24, 'Programming languages': 23,
        'Software engineering': 18, 'Algorithms & complexity': 147, 'Cryptography': 15, 'Logic & verification': 6,
        'Computational biology & bioinformatics': 0, 'Computer graphics': 4, 'Computer science education': 5, 'Economics & computation': 29,
        'Human-computer interaction': 206, 'Robotics': 66, 'Visualization': 7
    },
    'University of Washington': {
        'Artificial intelligence': 51, 'Computer vision': 115, 'Machine learning': 376, 'Natural language processing': 428,
        'The Web & information retrieval': 16, 'Computer architecture': 44, 'Computer networks': 81, 'Computer security': 60,
        'Databases': 50, 'Design automation': 5, 'Embedded & real-time systems': 2, 'High-performance computing': 0,
        'Mobile computing': 32, 'Measurement & performance analysis': 16, 'Operating systems': 38, 'Programming languages': 33,
        'Software engineering': 24, 'Algorithms & complexity': 97, 'Cryptography': 62, 'Logic & verification': 6,
        'Computational biology & bioinformatics': 6, 'Computer graphics': 33, 'Computer science education': 12, 'Economics & computation': 12,
        'Human-computer interaction': 452, 'Robotics': 103, 'Visualization': 19
    },
    'Stanford University': {
        'Artificial intelligence': 67, 'Computer vision': 336, 'Machine learning': 738, 'Natural language processing': 211,
        'The Web & information retrieval': 12, 'Computer architecture': 59, 'Computer networks': 49, 'Computer security': 46,
        'Databases': 33, 'Design automation': 23, 'Embedded & real-time systems': 0, 'High-performance computing': 18,
        'Mobile computing': 7, 'Measurement & performance analysis': 16, 'Operating systems': 17, 'Programming languages': 34,
        'Software engineering': 2, 'Algorithms & complexity': 159, 'Cryptography': 19, 'Logic & verification': 24,
        'Computational biology & bioinformatics': 5, 'Computer graphics': 100, 'Computer science education': 15, 'Economics & computation': 19,
        'Human-computer interaction': 162, 'Robotics': 223, 'Visualization': 11
    },
    'Cornell University': {
        'Artificial intelligence': 91, 'Computer vision': 153, 'Machine learning': 350, 'Natural language processing': 153,
        'The Web & information retrieval': 71, 'Computer architecture': 48, 'Computer networks': 41, 'Computer security': 113,
        'Databases': 46, 'Design automation': 36, 'Embedded & real-time systems': 1, 'High-performance computing': 4,
        'Mobile computing': 6, 'Measurement & performance analysis': 9, 'Operating systems': 16, 'Programming languages': 56,
        'Software engineering': 7, 'Algorithms & complexity': 53, 'Cryptography': 44, 'Logic & verification': 18,
        'Computational biology & bioinformatics': 1, 'Computer graphics': 48, 'Computer science education': 0, 'Economics & computation': 48,
        'Human-computer interaction': 135, 'Robotics': 89, 'Visualization': 5
    },
    'University of Maryland - College Park': {
        'Artificial intelligence': 184, 'Computer vision': 212, 'Machine learning': 343, 'Natural language processing': 167,
        'The Web & information retrieval': 20, 'Computer architecture': 3, 'Computer networks': 26, 'Computer security': 112,
        'Databases': 32, 'Design automation': 15, 'Embedded & real-time systems': 1, 'High-performance computing': 18,
        'Mobile computing': 19, 'Measurement & performance analysis': 19, 'Operating systems': 5, 'Programming languages': 20,
        'Software engineering': 9, 'Algorithms & complexity': 49, 'Cryptography': 37, 'Logic & verification': 0,
        'Computational biology & bioinformatics': 25, 'Computer graphics': 32, 'Computer science education': 8, 'Economics & computation': 18,
        'Human-computer interaction': 105, 'Robotics': 35, 'Visualization': 39
    },
    'Northeastern University': {
        'Artificial intelligence': 124, 'Computer vision': 108, 'Machine learning': 157, 'Natural language processing': 55,
        'The Web & information retrieval': 25, 'Computer architecture': 27, 'Computer networks': 7, 'Computer security': 123,
        'Databases': 41, 'Design automation': 44, 'Embedded & real-time systems': 4, 'High-performance computing': 30,
        'Mobile computing': 10, 'Measurement & performance analysis': 73, 'Operating systems': 6, 'Programming languages': 22,
        'Software engineering': 16, 'Algorithms & complexity': 68, 'Cryptography': 42, 'Logic & verification': 8,
        'Computational biology & bioinformatics': 9, 'Computer graphics': 2, 'Computer science education': 3, 'Economics & computation': 11,
        'Human-computer interaction': 141, 'Robotics': 75, 'Visualization': 31
    },
    'Purdue University': {
        'Artificial intelligence': 62, 'Computer vision': 38, 'Machine learning': 149, 'Natural language processing': 21,
        'The Web & information retrieval': 36, 'Computer architecture': 78, 'Computer networks': 18, 'Computer security': 182,
        'Databases': 34, 'Design automation': 4, 'Embedded & real-time systems': 4, 'High-performance computing': 20,
        'Mobile computing': 37, 'Measurement & performance analysis': 19, 'Operating systems': 15, 'Programming languages': 37,
        'Software engineering': 72, 'Algorithms & complexity': 35, 'Cryptography': 34, 'Logic & verification': 8,
        'Computational biology & bioinformatics': 10, 'Computer graphics': 17, 'Computer science education': 0, 'Economics & computation': 10,
        'Human-computer interaction': 24, 'Robotics': 33, 'Visualization': 13
    },
    'University of Wisconsin - Madison': {
        'Artificial intelligence': 49, 'Computer vision': 145, 'Machine learning': 330, 'Natural language processing': 24,
        'The Web & information retrieval': 6, 'Computer architecture': 62, 'Computer networks': 23, 'Computer security': 95,
        'Databases': 62, 'Design automation': 4, 'Embedded & real-time systems': 0, 'High-performance computing': 10,
        'Mobile computing': 19, 'Measurement & performance analysis': 19, 'Operating systems': 39, 'Programming languages': 49,
        'Software engineering': 18, 'Algorithms & complexity': 48, 'Cryptography': 10, 'Logic & verification': 26,
        'Computational biology & bioinformatics': 1, 'Computer graphics': 22, 'Computer science education': 4, 'Economics & computation': 0,
        'Human-computer interaction': 61, 'Robotics': 31, 'Visualization': 18
    },
    'University of Texas at Austin': {
        'Artificial intelligence': 85, 'Computer vision': 152, 'Machine learning': 364, 'Natural language processing': 110,
        'The Web & information retrieval': 3, 'Computer architecture': 32, 'Computer networks': 48, 'Computer security': 42,
        'Databases': 12, 'Design automation': 59, 'Embedded & real-time systems': 7, 'High-performance computing': 3,
        'Mobile computing': 17, 'Measurement & performance analysis': 9, 'Operating systems': 30, 'Programming languages': 43,
        'Software engineering': 13, 'Algorithms & complexity': 63, 'Cryptography': 45, 'Logic & verification': 11,
        'Computational biology & bioinformatics': 0, 'Computer graphics': 21, 'Computer science education': 1, 'Economics & computation': 4,
        'Human-computer interaction': 19, 'Robotics': 110, 'Visualization': 2
    },
    'University of Pennsylvania': {
        'Artificial intelligence': 51, 'Computer vision': 138, 'Machine learning': 331, 'Natural language processing': 172,
        'The Web & information retrieval': 5, 'Computer architecture': 34, 'Computer networks': 26, 'Computer security': 33,
        'Databases': 46, 'Design automation': 5, 'Embedded & real-time systems': 40, 'High-performance computing': 0,
        'Mobile computing': 4, 'Measurement & performance analysis': 5, 'Operating systems': 22, 'Programming languages': 54,
        'Software engineering': 7, 'Algorithms & complexity': 77, 'Cryptography': 8, 'Logic & verification': 22,
        'Computational biology & bioinformatics': 1, 'Computer graphics': 41, 'Computer science education': 1, 'Economics & computation': 27,
        'Human-computer interaction': 21, 'Robotics': 322, 'Visualization': 1
    },
    'Columbia University': {
        'Artificial intelligence': 42, 'Computer vision': 119, 'Machine learning': 220, 'Natural language processing': 84,
        'The Web & information retrieval': 14, 'Computer architecture': 24, 'Computer networks': 15, 'Computer security': 64,
        'Databases': 30, 'Design automation': 14, 'Embedded & real-time systems': 1, 'High-performance computing': 1,
        'Mobile computing': 20, 'Measurement & performance analysis': 25, 'Operating systems': 37, 'Programming languages': 10,
        'Software engineering': 25, 'Algorithms & complexity': 151, 'Cryptography': 17, 'Logic & verification': 3,
        'Computational biology & bioinformatics': 2, 'Computer graphics': 26, 'Computer science education': 3, 'Economics & computation': 22,
        'Human-computer interaction': 44, 'Robotics': 51, 'Visualization': 4
    },
    'Princeton University': {
        'Artificial intelligence': 16, 'Computer vision': 100, 'Machine learning': 377, 'Natural language processing': 61,
        'The Web & information retrieval': 11, 'Computer architecture': 64, 'Computer networks': 80, 'Computer security': 59,
        'Databases': 2, 'Design automation': 10, 'Embedded & real-time systems': 0, 'High-performance computing': 5,
        'Mobile computing': 25, 'Measurement & performance analysis': 15, 'Operating systems': 16, 'Programming languages': 27,
        'Software engineering': 2, 'Algorithms & complexity': 111, 'Cryptography': 8, 'Logic & verification': 10,
        'Computational biology & bioinformatics': 14, 'Computer graphics': 34, 'Computer science education': 1, 'Economics & computation': 24,
        'Human-computer interaction': 31, 'Robotics': 41, 'Visualization': 2
    },
    'New York University': {
        'Artificial intelligence': 43, 'Computer vision': 103, 'Machine learning': 380, 'Natural language processing': 131,
        'The Web & information retrieval': 23, 'Computer architecture': 12, 'Computer networks': 29, 'Computer security': 84,
        'Databases': 27, 'Design automation': 34, 'Embedded & real-time systems': 0, 'High-performance computing': 2,
        'Mobile computing': 1, 'Measurement & performance analysis': 11, 'Operating systems': 20, 'Programming languages': 16,
        'Software engineering': 3, 'Algorithms & complexity': 95, 'Cryptography': 55, 'Logic & verification': 8,
        'Computational biology & bioinformatics': 0, 'Computer graphics': 75, 'Computer science education': 3, 'Economics & computation': 4,
        'Human-computer interaction': 27, 'Robotics': 48, 'Visualization': 21
    },
    'University of California - Los Angeles': {
        'Artificial intelligence': 127, 'Computer vision': 79, 'Machine learning': 438, 'Natural language processing': 168,
        'The Web & information retrieval': 26, 'Computer architecture': 28, 'Computer networks': 46, 'Computer security': 12,
        'Databases': 8, 'Design automation': 45, 'Embedded & real-time systems': 2, 'High-performance computing': 2,
        'Mobile computing': 24, 'Measurement & performance analysis': 2, 'Operating systems': 15, 'Programming languages': 21,
        'Software engineering': 29, 'Algorithms & complexity': 40, 'Cryptography': 79, 'Logic & verification': 6,
        'Computational biology & bioinformatics': 31, 'Computer graphics': 6, 'Computer science education': 0, 'Economics & computation': 0,
        'Human-computer interaction': 13, 'Robotics': 6, 'Visualization': 0
    },
    'University of Massachusetts Amherst': {
        'Artificial intelligence': 99, 'Computer vision': 129, 'Machine learning': 241, 'Natural language processing': 117,
        'The Web & information retrieval': 40, 'Computer architecture': 4, 'Computer networks': 24, 'Computer security': 40,
        'Databases': 29, 'Design automation': 0, 'Embedded & real-time systems': 0, 'High-performance computing': 15,
        'Mobile computing': 57, 'Measurement & performance analysis': 26, 'Operating systems': 3, 'Programming languages': 8,
        'Software engineering': 37, 'Algorithms & complexity': 27, 'Cryptography': 0, 'Logic & verification': 0,
        'Computational biology & bioinformatics': 1, 'Computer graphics': 11, 'Computer science education': 2, 'Economics & computation': 3,
        'Human-computer interaction': 42, 'Robotics': 89, 'Visualization': 11
    },
    'University of Southern California': {
        'Artificial intelligence': 113, 'Computer vision': 82, 'Machine learning': 190, 'Natural language processing': 115,
        'The Web & information retrieval': 19, 'Computer architecture': 13, 'Computer networks': 41, 'Computer security': 13,
        'Databases': 26, 'Design automation': 5, 'Embedded & real-time systems': 7, 'High-performance computing': 9,
        'Mobile computing': 16, 'Measurement & performance analysis': 17, 'Operating systems': 5, 'Programming languages': 9,
        'Software engineering': 48, 'Algorithms & complexity': 21, 'Cryptography': 2, 'Logic & verification': 11,
        'Computational biology & bioinformatics': 0, 'Computer graphics': 10, 'Computer science education': 0, 'Economics & computation': 18,
        'Human-computer interaction': 13, 'Robotics': 121, 'Visualization': 0
    }
}

def generate_file_name():
    characters = string.ascii_letters + string.digits
    return ''.join(random.choices(characters, k=5)) + '.xlsx'

def calculate_affinity_scores(keywords, research_areas):
    keyword_embeddings = model.encode(keywords)
    area_embeddings = model.encode(research_areas)
    scores = []
    for area_embedding in area_embeddings:
        score = np.mean([np.dot(keyword_embedding, area_embedding) for keyword_embedding in keyword_embeddings])
        scores.append(score)
    return scores

def get_top_keywords(author_id, n=3):
    publications = client.search_notes(term=author_id, content='authors', group='NeurIPS.cc/2023/Conference', source='all')
    keywords = []
    for publication in publications:
        if 'keywords' in publication.content:
            keywords.extend(publication.content['keywords']['value'])
    counter = Counter(keywords)
    return [keyword for keyword, _ in counter.most_common(n)]

def get_author_profiles(submissions):
    author_profiles = {}
    for submission in submissions:
        author_ids = submission.content['authorids']['value']
        profiles = openreview.tools.get_profiles(client, author_ids)
        for author_id, profile in zip(author_ids, profiles):
            if author_id not in author_profiles and meets_conditions(profile):
                author_profiles[author_id] = profile
    return author_profiles

def meets_conditions(author_profile):
    history = author_profile.content.get('history', [])
    publications = client.search_notes(term=author_profile.id, content='authors', group='NeurIPS.cc/2023/Conference', source='all')

    has_professor = any('professor' in entry.get('position', '').lower() for entry in history)
    has_phd_2021 = any(
        'phd' in entry.get('position', '').lower() and
        str(entry.get('end', '')).startswith('2021')
        for entry in history if 'end' in entry
    )

    return (has_professor and has_phd_2021) and len(publications) >= 3

submissions = client.get_all_notes(content={'venueid': 'NeurIPS.cc/2023/Conference'})
author_profiles = get_author_profiles(submissions)

profiles_data = []
affinity_scores_data = []
similarity_scores = []

max_authors = 100
for author_id, author_profile in author_profiles.items():
    top_keywords = get_top_keywords(author_id)
    scores = calculate_affinity_scores(top_keywords, research_areas)
    
    profiles_data.append({
        'ID': author_id,
        'Affiliations': "; ".join({entry.get('institution', {}).get('name', 'N/A') for entry in author_profile.content.get('history', [])}),
        'History': "; ".join(["{} at {} ({}-{})".format(entry.get('position', 'N/A'), entry.get('institution', {}).get('name', 'N/A'), str(entry.get('start', 'N/A')), str(entry.get('end', 'N/A'))) for entry in author_profile.content.get('history', [])]),
        'Top Keywords': "; ".join(top_keywords)
    })

    score_dict = {'ID': author_id}
    score_dict.update(dict(zip(research_areas, scores)))
    affinity_scores_data.append(score_dict)

    # Calculate similarity for each university
    author_similarity = {'ID': author_id}
    for uni, papers in university_data.items():
        similarity_score = sum(score * papers.get(area, 0) for area, score in zip(research_areas, scores))
        total_papers = sum(papers.values())
        author_similarity[uni] = similarity_score / total_papers if total_papers else 0
    similarity_scores.append(author_similarity)

# Save to Excel with multiple sheets
file_name = generate_file_name()
wb = Workbook()
ws1 = wb.active
ws1.title = "Profiles"
for data in dataframe_to_rows(pd.DataFrame(profiles_data), index=False, header=True):
    ws1.append(data)

ws2 = wb.create_sheet(title="Affinity Scores")
for data in dataframe_to_rows(pd.DataFrame(affinity_scores_data), index=False, header=True):
    ws2.append(data)

ws3 = wb.create_sheet(title="Similarity Matrix")
similarity_df = pd.DataFrame(similarity_scores)
for row in dataframe_to_rows(similarity_df, index=False, header=True):
    ws3.append(row)

max_score = similarity_df.drop('ID', axis=1).max().max()
min_score = similarity_df.drop('ID', axis=1).min().min()

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=2, max_col=ws3.max_column):
    for cell in row:
        if cell.value is not None:
            # Calculate normalized score
            normalized_score = (cell.value - min_score) / (max_score - min_score)
            # Calculate red and green components
            red = int((1 - normalized_score) * 255)
            green = int(normalized_score * 255)
            cell.fill = PatternFill(start_color=f"{red:02X}{green:02X}00", fill_type="solid")
            cell.font = Font(color="FFFFFF")

# Determine unique most and least similar authors for each university
uni_priority = {uni: similarity_df[uni].max() - similarity_df[uni].min() for uni in university_data}
sorted_unis = sorted(uni_priority.items(), key=lambda x: x[1], reverse=True)

assigned_authors = set()
most_similar = {}
least_similar = {}
used_authors = set()

for uni, _ in sorted_unis:
    author_scores = similarity_df[['ID', uni]].sort_values(by=uni, ascending=False)
    
    # Most similar
    for _, row in author_scores.iterrows():
        if row['ID'] not in used_authors:
            most_similar[uni] = row['ID']
            used_authors.add(row['ID'])
            break

    # Least similar
    for _, row in author_scores[::-1].iterrows():
        if row['ID'] not in used_authors:
            least_similar[uni] = row['ID']
            used_authors.add(row['ID'])
            break

# Add tables for most and least similar authors
ws_most_similar = wb.create_sheet(title="Most Similar Authors")
ws_most_similar.append(['University', 'Most Similar Author'])
for uni, author in most_similar.items():
    ws_most_similar.append([uni, author])

ws_least_similar = wb.create_sheet(title="Least Similar Authors")
ws_least_similar.append(['University', 'Least Similar Author'])
for uni, author in least_similar.items():
    ws_least_similar.append([uni, author])

universities = list(university_data.keys())
authors = similarity_df['ID'].tolist()

# Create the cost matrix for the utilitarian approach
cost_matrix_utilitarian = -similarity_df[universities].to_numpy()

row_indices, col_indices = linear_sum_assignment(cost_matrix_utilitarian)

# Utilitarian Matching: find the combination of scores which avg to be the highest possible
cost_matrix_utilitarian = -similarity_df[universities].to_numpy()

row_indices, col_indices = linear_sum_assignment(cost_matrix_utilitarian)

utilitarian_matches = []
for row, col in zip(row_indices, col_indices):
    university = universities[col]
    author_id = similarity_df.iloc[row]['ID']
    score = -cost_matrix_utilitarian[row, col]  # Convert back to positive score for clarity
    utilitarian_matches.append({'University': university, 'Author ID': author_id, 'Score': score})

# Convert to DataFrame
utilitarian_df = pd.DataFrame(utilitarian_matches)
# Sorting by score to verify the optimality of the assignment
utilitarian_df.sort_values(by='Score', ascending=False, inplace=True)

ws_utilitarian = wb.create_sheet(title="Utilitarian Matching")
for data in dataframe_to_rows(utilitarian_df, index=False, header=True):
    ws_utilitarian.append(data)

# Egalitarian Matching: find the combination of scores with the lowest possible range
cost_matrix_egalitarian = similarity_df[universities].to_numpy()

# Pad the cost matrix with zeros if the number of authors is less than the number of universities
if len(authors) < len(universities):
    padding = np.zeros((len(universities) - len(authors), len(universities)))
    cost_matrix_egalitarian = np.vstack((cost_matrix_egalitarian, padding))

row_indices, col_indices = linear_sum_assignment(cost_matrix_egalitarian, maximize=True)

egalitarian_matches = []
for col, row in enumerate(row_indices):
    if row < len(authors):
        university = universities[col]
        author_id = authors[row]
        score = cost_matrix_egalitarian[row, col]
        egalitarian_matches.append({'University': university, 'Author ID': author_id, 'Score': score})
    else:
        university = universities[col]
        author_id = None
        score = None
        egalitarian_matches.append({'University': university, 'Author ID': author_id, 'Score': score})

# Convert to DataFrame
egalitarian_df = pd.DataFrame(egalitarian_matches)
ws_egalitarian = wb.create_sheet(title="Egalitarian Matching")
for data in dataframe_to_rows(egalitarian_df, index=False, header=True):
    ws_egalitarian.append(data)

wb.save(file_name)
print(f"Excel file has been created: {file_name}")